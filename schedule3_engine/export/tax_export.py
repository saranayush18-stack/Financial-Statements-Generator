"""
Tax exports.

1.  add_tax_sheets(wb, ...): appends 'IT Depreciation', 'CFL', 'Tax Comp',
    'Deferred Tax', 'Tax Audit' sheets to the financial-statements workbook,
    with LIVE cross-sheet formulas exactly in Book1's pattern:
        Tax Comp!  depreciation add-back  -> book figure
        Tax Comp!  IT depreciation        -> ='IT Depreciation'!<total dep>
        Tax Comp!  b/f loss set-off       -> =MIN(income, 'CFL'!...)
        Deferred Tax! Fixed assets        -> ='IT Depreciation'!<closing WDV>
2.  build_computation_pdf(...): a "paper return"-style Computation of Total
    Income and Tax statement (the working that accompanies an ITR filing),
    plus the tax-audit applicability annexure.
"""
from __future__ import annotations

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

from models import Company
from core.income_computation import TaxComputation
from core.it_depreciation import ITDepreciationSchedule
from core.deferred_tax import DeferredTaxComputation
from core.tax_audit import AuditApplicability, ClauseItem
from data.tax_config import EntityType, TaxRegime

TNR = "Times New Roman"
ACCT_FMT = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
PCT_FMT = "0.00%"
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
TOTAL_FILL = PatternFill("solid", fgColor="BDD7EE")
OVERRIDE_FILL = PatternFill("solid", fgColor="FFF2CC")   # amber = manually overridden
S_MEDIUM = Side(style="medium", color="000000")
S_THIN = Side(style="thin", color="000000")
S_DOUBLE = Side(style="double", color="000000")


def _tnr(size=11, bold=False, italic=False, color=None):
    return Font(name=TNR, size=size, bold=bold, italic=italic, color=color)


def _box(ws, r1, c1, r2, c2, style="thin"):
    side = {"thin": S_THIN, "medium": S_MEDIUM, "double": S_DOUBLE}[style]
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=col)
            b = cell.border
            cell.border = Border(
                top=side if row == r1 else b.top,
                bottom=side if row == r2 else b.bottom,
                left=side if col == c1 else b.left,
                right=side if col == c2 else b.right,
            )


def _fit(ws, last_col):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:{get_column_letter(last_col)}{ws.max_row}"


def _sheet_header(ws, company: Company, title: str, last_col: int) -> int:
    r = 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=last_col)
    c = ws.cell(row=r, column=2, value=company.name)
    c.font = _tnr(12, bold=True)
    c.alignment = Alignment(horizontal="center")
    r += 1
    if company.pan:
        ws.cell(row=r, column=2, value=f"PAN: {company.pan}").font = _tnr(10)
    ws.cell(row=r, column=last_col - 1, value="Tax Year: 2026-27").font = _tnr(10)
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=last_col)
    c = ws.cell(row=r, column=2, value=title)
    c.font = _tnr(11, bold=True)
    c.alignment = Alignment(horizontal="center")
    return r + 2


# ---------------------------------------------------------------------------
# 1. IT Depreciation sheet (must be written BEFORE Tax Comp so cell refs exist)
# ---------------------------------------------------------------------------
def _write_it_depreciation_sheet(wb, company: Company, sched: ITDepreciationSchedule) -> dict:
    ws = wb.create_sheet("IT Depreciation")
    ws.sheet_view.showGridLines = False
    r = _sheet_header(ws, company, "Depreciation as per the Income-tax Act, 2025 (Block of Assets - WDV)", 9)

    hdr = ["Sl.", "Block of Assets", "Rate", "WDV as on\n01-04-2026",
           "Additions\n(>=180 days)", "Additions\n(<180 days)", "Deletions", "Depreciation", "WDV as on\n31-03-2027"]
    hdr_row = r
    for c, h in enumerate(hdr, start=2):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = _tnr(10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 34
    _box(ws, r, 2, r, 10, "medium")
    r += 1

    first_data = r
    for i, row in enumerate(sched.rows, start=1):
        ws.cell(row=r, column=2, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=row.block_name).font = _tnr(10)
        ws.cell(row=r, column=4, value=row.rate).number_format = PCT_FMT
        ws.cell(row=r, column=5, value=row.opening_wdv).number_format = ACCT_FMT
        ws.cell(row=r, column=6, value=row.additions_180_plus).number_format = ACCT_FMT
        ws.cell(row=r, column=7, value=row.additions_less_180).number_format = ACCT_FMT
        ws.cell(row=r, column=8, value=row.deletions).number_format = ACCT_FMT
        # Depreciation = MIN((open+add180-del)*rate + add<180*rate/2, base) floored at 0
        ws.cell(row=r, column=9,
                value=f"=MAX(MIN((E{r}+F{r}-H{r})*D{r}+G{r}*D{r}/2,E{r}+F{r}+G{r}-H{r}),0)"
                ).number_format = ACCT_FMT
        ws.cell(row=r, column=10, value=f"=E{r}+F{r}+G{r}-H{r}-I{r}").number_format = ACCT_FMT
        for c in range(2, 11):
            cell = ws.cell(row=r, column=c)
            cell.font = _tnr(10)
        r += 1
    last_data = r - 1

    ws.cell(row=r, column=3, value="Grand Total").font = _tnr(10, bold=True)
    for col, letter in [(5, "E"), (6, "F"), (7, "G"), (8, "H"), (9, "I"), (10, "J")]:
        if last_data >= first_data:
            ws.cell(row=r, column=col, value=f"=SUM({letter}{first_data}:{letter}{last_data})")
        else:
            ws.cell(row=r, column=col, value=0)
        ws.cell(row=r, column=col).number_format = ACCT_FMT
        ws.cell(row=r, column=col).font = _tnr(10, bold=True)
        ws.cell(row=r, column=col).fill = TOTAL_FILL
    _box(ws, r, 5, r, 10, "double")
    total_row = r

    for col, w in {1: 3, 2: 5, 3: 34, 4: 9, 5: 16, 6: 15, 7: 15, 8: 13, 9: 15, 10: 16}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    _fit(ws, 10)
    return {"dep_total": f"'IT Depreciation'!I{total_row}",
            "closing_wdv_total": f"'IT Depreciation'!J{total_row}"}


# ---------------------------------------------------------------------------
# 2. Carried Forward Losses (CFL) sheet
# ---------------------------------------------------------------------------
def _write_cfl_sheet(wb, company: Company, bf_loss: float, unabs_dep: float) -> dict:
    ws = wb.create_sheet("CFL")
    ws.sheet_view.showGridLines = False
    r = _sheet_header(ws, company, "Business Loss and Unabsorbed Depreciation Schedule", 7)
    hdr = ["Particulars", "Brought Forward", "Set off in Current Year", "Carried Forward"]
    hdr_row = r
    for c, h in enumerate(hdr, start=2):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = _tnr(10, bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    _box(ws, r, 2, r, 5, "medium")
    r += 1
    loss_row = r
    ws.cell(row=r, column=2, value="Brought forward business loss").font = _tnr(10)
    ws.cell(row=r, column=3, value=bf_loss).number_format = ACCT_FMT
    ws.cell(row=r, column=4, value="='Tax Comp'!__LOSS_SETOFF__").number_format = ACCT_FMT
    ws.cell(row=r, column=5, value=f"=C{r}-D{r}").number_format = ACCT_FMT
    r += 1
    dep_row = r
    ws.cell(row=r, column=2, value="Brought forward unabsorbed depreciation").font = _tnr(10)
    ws.cell(row=r, column=3, value=unabs_dep).number_format = ACCT_FMT
    ws.cell(row=r, column=4, value="='Tax Comp'!__UNABS_SETOFF__").number_format = ACCT_FMT
    ws.cell(row=r, column=5, value=f"=C{r}-D{r}").number_format = ACCT_FMT
    for col, w in {1: 3, 2: 42, 3: 18, 4: 20, 5: 18}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    _fit(ws, 5)
    return {"bf_loss_cell": f"'CFL'!C{loss_row}", "unabs_dep_cell": f"'CFL'!C{dep_row}",
            "loss_setoff_target": (loss_row, 4), "unabs_setoff_target": (dep_row, 4)}


# ---------------------------------------------------------------------------
# 3. Tax Comp sheet (Book1 pattern, formula-linked)
# ---------------------------------------------------------------------------
def _write_tax_comp_sheet(wb, company: Company, comp: TaxComputation,
                          dep_refs: dict, cfl_refs: dict) -> dict:
    ws = wb.create_sheet("Tax Comp")
    ws.sheet_view.showGridLines = False
    r = _sheet_header(ws, company, "COMPUTATION OF TOTAL INCOME AND TAX", 7)
    ws.cell(row=r, column=2, value=f"Status: {comp.entity_type.value}").font = _tnr(10)
    r += 1
    ws.cell(row=r, column=2, value=f"Regime: {comp.regime.value}").font = _tnr(10)
    r += 2

    def line(label, col_e=None, col_f=None, bold=False, italic=False, fill=None,
             fmt=ACCT_FMT, indent=0):
        nonlocal r
        c = ws.cell(row=r, column=3, value=("     " * indent) + label)
        c.font = _tnr(10, bold=bold, italic=italic)
        if col_e is not None:
            ce = ws.cell(row=r, column=5, value=col_e)
            ce.number_format = fmt
            ce.font = _tnr(10, bold=bold)
            if fill:
                ce.fill = fill
        if col_f is not None:
            cf = ws.cell(row=r, column=6, value=col_f)
            cf.number_format = fmt
            cf.font = _tnr(10, bold=bold)
            if fill:
                cf.fill = fill
        r += 1

    line("Income from Business or Profession", bold=True)
    line("Net Profit as per Statement of Profit and Loss", col_f=comp.net_profit_as_per_pnl)
    pbt_row = r - 1

    line("Add: Inadmissible expenses", bold=True)
    add_first = r
    for adj in comp.additions:
        ws.cell(row=r, column=3, value="     " + adj.label).font = _tnr(10)
        ws.cell(row=r, column=4, value=adj.section_ref).font = _tnr(8, italic=True)
        ce = ws.cell(row=r, column=5, value=adj.final)
        ce.number_format = ACCT_FMT
        if adj.is_overridden:
            ce.fill = OVERRIDE_FILL
            ws.cell(row=r, column=7, value=f"manual (auto: {adj.auto:,.2f})").font = _tnr(8, italic=True, color="808080")
        r += 1
    add_last = r - 1
    line("Total additions", col_f=f"=SUM(E{add_first}:E{add_last})", bold=True)
    add_total_row = r - 1

    line("Less: Admissible expenses", bold=True)
    ded_first = r
    for adj in comp.deductions:
        ws.cell(row=r, column=3, value="     " + adj.label).font = _tnr(10)
        ws.cell(row=r, column=4, value=adj.section_ref).font = _tnr(8, italic=True)
        if adj.code == "LESS_IT_DEP" and not adj.is_overridden:
            ce = ws.cell(row=r, column=5, value=f"=+{dep_refs['dep_total']}")
        else:
            ce = ws.cell(row=r, column=5, value=adj.final)
            if adj.is_overridden:
                ce.fill = OVERRIDE_FILL
                ws.cell(row=r, column=7, value=f"manual (auto: {adj.auto:,.2f})").font = _tnr(8, italic=True, color="808080")
        ce.number_format = ACCT_FMT
        r += 1
    ded_last = r - 1
    line("Total deductions", col_f=f"=SUM(E{ded_first}:E{ded_last})", bold=True)
    ded_total_row = r - 1

    line("Total Business Income", col_f=f"=F{pbt_row}+F{add_total_row}-F{ded_total_row}", bold=True,
         fill=TOTAL_FILL)
    biz_income_row = r - 1

    line("Less: Set-off of brought forward business loss",
         col_f=f"=MIN(MAX(F{biz_income_row},0),{cfl_refs['bf_loss_cell']})")
    loss_setoff_row = r - 1
    line("Less: Set-off of unabsorbed depreciation",
         col_f=f"=MIN(MAX(F{biz_income_row}-F{loss_setoff_row},0),{cfl_refs['unabs_dep_cell']})")
    unabs_setoff_row = r - 1
    line("Taxable Total Income", col_f=f"=F{biz_income_row}-F{loss_setoff_row}-F{unabs_setoff_row}",
         bold=True, fill=TOTAL_FILL)
    ti_row = r - 1
    _box(ws, ti_row, 6, ti_row, 6, "double")
    r += 1

    # Tax block -- rate as a visible, editable cell (H column) like Book1
    base_rate = comp.tax_before_surcharge / comp.taxable_income if comp.taxable_income > 0 else 0.0
    ws.cell(row=r, column=8, value=comp.surcharge_rate).number_format = PCT_FMT
    surcharge_rate_cell = f"H{r}"
    if comp.entity_type == EntityType.COMPANY or comp.entity_type == EntityType.FIRM_LLP:
        ws.cell(row=r, column=7, value=base_rate if comp.taxable_income > 0 else 0.30).number_format = PCT_FMT
        rate_cell = f"G{r}"
        line(f"Tax on total income @ {base_rate:.0%}" if comp.taxable_income > 0 else "Tax on total income",
             col_f=f"=IF(F{ti_row}>0,ROUND(F{ti_row}*{rate_cell},0),0)")
    else:
        # slab computation embedded as value (slabs don't collapse to one rate)
        line("Tax on total income (slab-based)", col_f=comp.tax_before_surcharge)
    tax_row = r - 1
    line(f"Add: Surcharge @ {comp.surcharge_rate:.0%} (marginal relief applied)",
         col_f=comp.surcharge)
    sur_row = r - 1
    line("Add: Health & Education Cess @ 4%", col_f=f"=ROUND((F{tax_row}+F{sur_row})*0.04,0)")
    cess_row = r - 1
    line("Total Tax Liability", col_f=f"=F{tax_row}+F{sur_row}+F{cess_row}", bold=True, fill=TOTAL_FILL)
    liability_row = r - 1
    r += 1

    line("Less: Prepaid taxes", bold=True)
    line("Tax Deducted at Source", col_e=comp.tds_credit, indent=1)
    tds_row = r - 1
    line("Tax Collected at Source", col_e=comp.tcs_credit, indent=1)
    line("Advance Tax paid", col_e=comp.advance_tax_paid, indent=1)
    line("Self-Assessment Tax paid", col_e=comp.self_assessment_tax_paid, indent=1)
    prepaid_last = r - 1
    line("Total prepaid", col_f=f"=SUM(E{tds_row}:E{prepaid_last})", bold=True)
    prepaid_row = r - 1

    line("Add: Interest for defaults", bold=True)
    line("Late filing of return (old 234A pattern)", col_e=comp.interest_234a.final, indent=1)
    int_first = r - 1
    line("Advance tax default (old 234B pattern)", col_e=comp.interest_234b.final, indent=1)
    line("Advance tax deferment (old 234C pattern)", col_e=comp.interest_234c.final, indent=1)
    int_last = r - 1
    line("Total interest", col_f=f"=SUM(E{int_first}:E{int_last})", bold=True)
    int_row = r - 1

    line("NET TAX PAYABLE / (REFUND)",
         col_f=f"=ROUND(F{liability_row}-F{prepaid_row}+F{int_row},-1)", bold=True, fill=TOTAL_FILL)
    _box(ws, r - 1, 6, r - 1, 6, "double")

    for col, w in {1: 3, 2: 6, 3: 52, 4: 26, 5: 16, 6: 18, 7: 20, 8: 9}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    _fit(ws, 8)
    return {"loss_setoff_cell": f"F{loss_setoff_row}", "unabs_setoff_cell": f"F{unabs_setoff_row}"}


def _patch_cfl_formulas(wb, cfl_refs: dict, tax_refs: dict):
    ws = wb["CFL"]
    lr, lc = cfl_refs["loss_setoff_target"]
    ws.cell(row=lr, column=lc, value=f"='Tax Comp'!{tax_refs['loss_setoff_cell']}")
    ur, uc = cfl_refs["unabs_setoff_target"]
    ws.cell(row=ur, column=uc, value=f"='Tax Comp'!{tax_refs['unabs_setoff_cell']}")


# ---------------------------------------------------------------------------
# 4. Deferred Tax sheet (Book1 pattern)
# ---------------------------------------------------------------------------
def _write_deferred_tax_sheet(wb, company: Company, dt: DeferredTaxComputation,
                              dep_refs: dict, book_wdv: float):
    ws = wb.create_sheet("Deferred Tax")
    ws.sheet_view.showGridLines = False
    r = _sheet_header(ws, company, "Deferred Tax Working (AS 22 pattern)", 8)
    hdr = ["Particulars", "Opening DTA/(DTL)", "As per Income-tax", "As per Books",
           "Timing Difference", "Closing DTA/(DTL)", "Charge/(Credit) for Year"]
    for c, h in enumerate(hdr, start=2):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = _tnr(10, bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[r].height = 30
    _box(ws, r, 2, r, 8, "medium")
    rate_cell_row = r
    ws.cell(row=r, column=9, value=dt.effective_rate).number_format = "0.000%"
    ws.cell(row=r, column=9).font = _tnr(9, italic=True)
    r += 1

    first = r
    for i, item in enumerate(dt.items):
        ws.cell(row=r, column=2, value=item.particulars).font = _tnr(10)
        ws.cell(row=r, column=3, value=item.opening_dta_dtl).number_format = ACCT_FMT
        if i == 0:  # Fixed assets: live link to IT Depreciation closing WDV
            ws.cell(row=r, column=4, value=f"=+{dep_refs['closing_wdv_total']}").number_format = ACCT_FMT
            ws.cell(row=r, column=5, value=book_wdv).number_format = ACCT_FMT
        else:
            ws.cell(row=r, column=4, value=item.as_per_income_tax).number_format = ACCT_FMT
            ws.cell(row=r, column=5, value=item.as_per_books).number_format = ACCT_FMT
        if item.override_timing_difference is not None:
            ce = ws.cell(row=r, column=6, value=item.override_timing_difference)
            ce.fill = OVERRIDE_FILL
        else:
            ws.cell(row=r, column=6, value=f"=E{r}-D{r}")
        ws.cell(row=r, column=6).number_format = ACCT_FMT
        ws.cell(row=r, column=7, value=f"=F{r}*$I${rate_cell_row}").number_format = ACCT_FMT
        ws.cell(row=r, column=8, value=f"=G{r}-C{r}").number_format = ACCT_FMT
        r += 1
    last = r - 1

    ws.cell(row=r, column=2, value="Total").font = _tnr(10, bold=True)
    for col, letter in [(3, "C"), (6, "F"), (7, "G"), (8, "H")]:
        ws.cell(row=r, column=col, value=f"=SUM({letter}{first}:{letter}{last})").number_format = ACCT_FMT
        ws.cell(row=r, column=col).font = _tnr(10, bold=True)
        ws.cell(row=r, column=col).fill = TOTAL_FILL
    _box(ws, r, 3, r, 8, "double")
    r += 2
    ws.cell(row=r, column=2, value=(
        "Positive closing = Deferred Tax Asset; negative = Deferred Tax Liability. "
        "Charge/(Credit) for the year flows to the Statement of Profit and Loss as "
        "deferred tax and to the Balance Sheet as DTA/DTL."
    )).font = _tnr(9, italic=True, color="808080")
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)

    for col, w in {1: 3, 2: 44, 3: 17, 4: 17, 5: 17, 6: 17, 7: 17, 8: 20, 9: 10}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    _fit(ws, 9)


# ---------------------------------------------------------------------------
# 5. Tax Audit sheet (applicability + Form 26 checklist)
# ---------------------------------------------------------------------------
def _write_tax_audit_sheet(wb, company: Company, applicability: AuditApplicability,
                           checklist: list[ClauseItem]):
    ws = wb.create_sheet("Tax Audit")
    ws.sheet_view.showGridLines = False
    r = _sheet_header(ws, company, "Tax Audit u/s 63, Income-tax Act, 2025 — Form No. 26 Working", 8)

    verdict = "TAX AUDIT REQUIRED" if applicability.required else "Tax audit NOT required"
    c = ws.cell(row=r, column=2, value=verdict)
    c.font = _tnr(12, bold=True, color="9C0006" if applicability.required else "006100")
    r += 1
    for reason in applicability.reasons:
        ws.cell(row=r, column=2, value="• " + reason).font = _tnr(10)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        r += 1
    if applicability.required:
        ws.cell(row=r, column=2, value=(
            f"Report: {applicability.form}. {applicability.due_note} "
            f"Fee for default u/s 446: Rs. {applicability.fee_if_defaulted:,.0f} "
            f"(0.5% of turnover, max Rs. 1,50,000)."
        )).font = _tnr(9, italic=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        r += 1
    r += 1

    hdr = ["Cl.", "Particulars (Form 26, Part B)", "Old 3CD ref", "Engine pre-fill",
           "Auditor response / particulars", "Status", "Guidance"]
    for c_i, h in enumerate(hdr, start=2):
        cell = ws.cell(row=r, column=c_i, value=h)
        cell.font = _tnr(10, bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    _box(ws, r, 2, r, 8, "medium")
    r += 1
    for item in checklist:
        ws.cell(row=r, column=2, value=item.no).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=r, column=3, value=item.title).font = _tnr(9)
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=4, value=item.old_3cd_ref).font = _tnr(9)
        ws.cell(row=r, column=5, value=item.auto_value).font = _tnr(9)
        ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=6, value=item.response).font = _tnr(9)
        st_cell = ws.cell(row=r, column=7, value=item.status)
        st_cell.font = _tnr(9, bold=True,
                            color="006100" if item.status == "Completed" else "9C6500")
        ws.cell(row=r, column=8, value=item.guidance).font = _tnr(8, italic=True, color="808080")
        ws.cell(row=r, column=8).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    for col, w in {1: 3, 2: 5, 3: 52, 4: 12, 5: 26, 6: 30, 7: 12, 8: 40}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    _fit(ws, 8)
    ws.cell(row=r + 1, column=2, value=(
        "Clause set per Form 26 (Rule 47, Income-tax Rules, 2026) as publicly notified; "
        "anchor clauses 49-51 (TDS/TCS) confirmed. Where the CBDT utility's final numbering "
        "differs, update data/form26_clauses.py."
    )).font = _tnr(8, italic=True, color="808080")


# ---------------------------------------------------------------------------
# PUBLIC: add all tax sheets
# ---------------------------------------------------------------------------
def add_tax_sheets(wb, company: Company, comp: TaxComputation,
                   it_dep: ITDepreciationSchedule, dt: DeferredTaxComputation,
                   book_closing_wdv: float,
                   applicability: AuditApplicability, checklist: list[ClauseItem]):
    dep_refs = _write_it_depreciation_sheet(wb, company, it_dep)
    cfl_refs = _write_cfl_sheet(wb, company, comp.brought_forward_business_loss,
                                comp.unabsorbed_depreciation_bf)
    tax_refs = _write_tax_comp_sheet(wb, company, comp, dep_refs, cfl_refs)
    _patch_cfl_formulas(wb, cfl_refs, tax_refs)
    _write_deferred_tax_sheet(wb, company, dt, dep_refs, book_closing_wdv)
    _write_tax_audit_sheet(wb, company, applicability, checklist)


# ---------------------------------------------------------------------------
# PAPER RETURN -- Computation of Total Income and Tax (PDF)
# ---------------------------------------------------------------------------
def build_computation_pdf(output_path: str, company: Company, comp: TaxComputation,
                          it_dep: ITDepreciationSchedule,
                          dt: DeferredTaxComputation,
                          applicability: AuditApplicability) -> str:
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("H", parent=styles["Title"], fontName="Times-Bold",
                               fontSize=13, alignment=TA_CENTER, spaceAfter=2))
    styles.add(ParagraphStyle("Sub", parent=styles["Normal"], fontName="Times-Roman",
                               fontSize=9, alignment=TA_CENTER, spaceAfter=2))
    styles.add(ParagraphStyle("B", parent=styles["Normal"], fontName="Times-Roman",
                               fontSize=9, leading=12))

    def fmt(n):
        if n is None or abs(n) < 0.005:
            return "-"
        s = f"{abs(n):,.2f}"
        return f"({s})" if n < 0 else s

    doc = SimpleDocTemplate(output_path, pagesize=A4, topMargin=16 * mm,
                            bottomMargin=16 * mm, leftMargin=18 * mm, rightMargin=18 * mm,
                            title=f"{company.name} - Computation of Income - Tax Year 2026-27")
    story = []
    story.append(Paragraph(company.name, styles["H"]))
    meta = []
    if company.pan:
        meta.append(f"PAN: {company.pan}")
    meta.append(f"Status: {comp.entity_type.value}")
    meta.append("Tax Year: 2026-27 (Income-tax Act, 2025)")
    meta.append(f"Regime: {comp.regime.value}")
    story.append(Paragraph(" &nbsp;|&nbsp; ".join(meta), styles["Sub"]))
    story.append(Paragraph("COMPUTATION OF TOTAL INCOME AND TAX", styles["H"]))
    story.append(Spacer(1, 6))

    rows = [["Particulars", "Rs.", "Rs."]]
    bold_rows = {0}
    rows.append(["Income from Business or Profession", "", ""]); bold_rows.add(len(rows) - 1)
    rows.append(["Net Profit as per Statement of Profit and Loss", "", fmt(comp.net_profit_as_per_pnl)])
    rows.append(["Add: Inadmissible expenses", "", ""]); bold_rows.add(len(rows) - 1)
    for adj in comp.additions:
        marker = " *" if adj.is_overridden else ""
        rows.append([f"    {adj.label}{marker}", fmt(adj.final), ""])
    rows.append(["Total additions", "", fmt(comp.total_additions)]); bold_rows.add(len(rows) - 1)
    rows.append(["Less: Admissible expenses", "", ""]); bold_rows.add(len(rows) - 1)
    for adj in comp.deductions:
        marker = " *" if adj.is_overridden else ""
        rows.append([f"    {adj.label}{marker}", fmt(adj.final), ""])
    rows.append(["Total deductions", "", fmt(comp.total_deductions)]); bold_rows.add(len(rows) - 1)
    rows.append(["Total Business Income", "", fmt(comp.business_income)]); bold_rows.add(len(rows) - 1)
    rows.append(["Less: Set-off of brought forward business loss", "", fmt(comp.loss_set_off)])
    rows.append(["Less: Set-off of unabsorbed depreciation", "", fmt(comp.unabsorbed_dep_set_off)])
    rows.append(["TAXABLE TOTAL INCOME", "", fmt(comp.taxable_income)]); bold_rows.add(len(rows) - 1)
    rows.append(["Tax on total income", "", fmt(comp.tax_before_surcharge)])
    rows.append([f"Add: Surcharge @ {comp.surcharge_rate:.0%} (with marginal relief)", "", fmt(comp.surcharge)])
    rows.append(["Add: Health & Education Cess @ 4%", "", fmt(comp.cess)])
    rows.append(["Total Tax Liability", "", fmt(comp.total_tax_liability)]); bold_rows.add(len(rows) - 1)
    rows.append(["Less: TDS / TCS credits", fmt(comp.tds_credit + comp.tcs_credit), ""])
    rows.append(["Less: Advance tax paid", fmt(comp.advance_tax_paid), ""])
    rows.append(["Less: Self-assessment tax paid", fmt(comp.self_assessment_tax_paid), ""])
    rows.append(["Total prepaid taxes", "", fmt(comp.prepaid_taxes)]); bold_rows.add(len(rows) - 1)
    rows.append(["Add: Interest for defaults (old 234A/B/C pattern)", "", fmt(comp.total_interest)])
    rows.append(["NET TAX PAYABLE / (REFUND) — rounded to Rs. 10", "", fmt(comp.net_payable)])
    bold_rows.add(len(rows) - 1)

    cmds = [
        ("FONTNAME", (0, 0), (-1, -1), "Times-Roman"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
        ("BOX", (0, 0), (-1, 0), 1, colors.black),
        ("LINEBELOW", (0, len(rows) - 1), (-1, len(rows) - 1), 1.75, colors.black),
    ]
    for i in bold_rows:
        cmds.append(("FONTNAME", (0, i), (-1, i), "Times-Bold"))
    t = Table(rows, colWidths=[112 * mm, 30 * mm, 30 * mm])
    t.setStyle(TableStyle(cmds))
    story.append(t)
    story.append(Spacer(1, 4))
    if any(a.is_overridden for a in comp.additions + comp.deductions):
        story.append(Paragraph("* manually overridden figure (auto value replaced by preparer)", styles["B"]))
    story.append(Paragraph(
        "Prepared under the Income-tax Act, 2025 for Tax Year 2026-27. Verify rates, "
        "surcharge marginal relief and interest against the enacted Finance Act, 2026 "
        "before filing.", styles["B"]))

    # Annexure: IT depreciation summary + deferred tax + audit applicability
    story.append(PageBreak())
    story.append(Paragraph("Annexure 1 — Depreciation as per the Income-tax Act (Block WDV)", styles["H"]))
    drows = [["Block", "Rate", "Opening WDV", "Add (>=180d)", "Add (<180d)", "Depreciation", "Closing WDV"]]
    for b in it_dep.rows:
        drows.append([b.block_name, f"{b.rate:.0%}", fmt(b.opening_wdv), fmt(b.additions_180_plus),
                      fmt(b.additions_less_180), fmt(b.depreciation), fmt(b.closing_wdv)])
    drows.append(["Total", "", fmt(it_dep.total_opening), fmt(it_dep.total_additions_180_plus),
                  fmt(it_dep.total_additions_less_180), fmt(it_dep.total_depreciation),
                  fmt(it_dep.total_closing_wdv)])
    dt_t = Table(drows, colWidths=[52 * mm, 12 * mm, 22 * mm, 22 * mm, 21 * mm, 22 * mm, 22 * mm])
    dt_t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Times-Roman"), ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Times-Bold"),
        ("FONTNAME", (0, len(drows) - 1), (-1, len(drows) - 1), "Times-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"), ("BOX", (0, 0), (-1, -1), 0.75, colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.grey),
    ]))
    story.append(dt_t)
    story.append(Spacer(1, 10))

    story.append(Paragraph("Annexure 2 — Deferred Tax Working", styles["H"]))
    dtr = [["Particulars", "As per IT", "As per Books", "Timing Diff.", "Closing DTA/(DTL)"]]
    for item in dt.items:
        dtr.append([item.particulars, fmt(item.as_per_income_tax), fmt(item.as_per_books),
                    fmt(item.timing_difference), fmt(dt.closing_for(item))])
    dtr.append(["Total (movement for year: " + fmt(dt.total_movement) + ")", "", "", "", fmt(dt.total_closing)])
    dt2 = Table(dtr, colWidths=[70 * mm, 26 * mm, 26 * mm, 26 * mm, 26 * mm])
    dt2.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Times-Roman"), ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Times-Bold"),
        ("FONTNAME", (0, len(dtr) - 1), (-1, len(dtr) - 1), "Times-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"), ("BOX", (0, 0), (-1, -1), 0.75, colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.grey),
    ]))
    story.append(dt2)
    story.append(Spacer(1, 10))

    story.append(Paragraph("Annexure 3 — Tax Audit Applicability (s.63, IT Act 2025)", styles["H"]))
    verdict = ("TAX AUDIT REQUIRED — report in " + applicability.form
               if applicability.required else "Tax audit not required on stated facts.")
    story.append(Paragraph(verdict, styles["B"]))
    for reason in applicability.reasons:
        story.append(Paragraph("• " + reason, styles["B"]))
    if applicability.required:
        story.append(Paragraph(applicability.due_note, styles["B"]))

    doc.build(story)
    return output_path
