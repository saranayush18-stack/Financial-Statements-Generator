"""
Excel Export.

Produces a single professionally formatted workbook with sheets:
TB | Mapping | Balance Sheet | P&L | Cash Flow | Notes | Validation | Ratios

Uses openpyxl. Currency cells use formulas referencing the TB sheet wherever
practical (Balance Sheet / P&L totals sum the Notes sheet ranges) so the
workbook recalculates if a ledger amount is corrected -- it is not just a
snapshot of Python-computed numbers.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import Company, TrialBalance, MappingEntry
from core.statement_generator import BalanceSheet, ProfitAndLoss, CashFlowStatement
from core.notes_generator import Note
from core.ratios import RatioResult
from models import ValidationIssue
from core.soce_generator import StatementOfChangesInEquity
from core.ageing import AgeingGrid, BUCKET_LABELS, CATEGORY_LABELS

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="BDD7EE")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT = '#,##0.00;(#,##0.00);"-"'

# --- Schedule III statutory-format constants (BS / P&L / Notes), matching the
# look of a standard signed financial statement: Times New Roman, accounting
# number format with bracketed negatives, medium box round headers, double
# rule under grand totals. ---
TNR = "Times New Roman"
ACCT_FMT = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
S_THIN = Side(style="thin", color="000000")
S_MEDIUM = Side(style="medium", color="000000")
S_DOUBLE = Side(style="double", color="000000")


def _tnr(size=11, bold=False, italic=False, color=None):
    return Font(name=TNR, size=size, bold=bold, italic=italic, color=color)


def _fit_to_page(ws, last_col: int = 8):
    """Landscape + fit-to-width print setup so amount columns never fall off
    the printable page (portrait with 8 columns clips columns G/H)."""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_area = f"A1:{get_column_letter(last_col)}{ws.max_row}"


def _box_border(ws, r1, c1, r2, c2, style="thin"):
    """Draw a border box (given side style) around the rectangular range."""
    side = S_THIN if style == "thin" else (S_MEDIUM if style == "medium" else S_DOUBLE)
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=col)
            b = cell.border
            top = side if row == r1 else b.top
            bottom = side if row == r2 else b.bottom
            left = side if col == c1 else b.left
            right = side if col == c2 else b.right
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)


def _note_number(note_ref: str | None) -> str:
    import re
    if not note_ref:
        return ""
    m = re.match(r"Note\s+(\d+)", note_ref)
    return m.group(1) if m else ""


def _write_schedule3_header(ws, company: Company, statement_title: str, fy_label: str, last_col: int = 8):
    """Company name / CIN / address / statement title block, matching a
    statutory signed financial statement layout. Returns the next free row."""
    r = 2
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=last_col)
    c = ws.cell(row=r, column=2, value=company.name)
    c.font = _tnr(12, bold=True)
    c.alignment = Alignment(horizontal="center")
    r += 1
    if company.cin:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=last_col)
        c = ws.cell(row=r, column=2, value=f"CIN: {company.cin}")
        c.font = _tnr(10)
        c.alignment = Alignment(horizontal="center")
        r += 1
    if company.registered_office:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=last_col)
        c = ws.cell(row=r, column=2, value=company.registered_office)
        c.font = _tnr(10)
        c.alignment = Alignment(horizontal="center")
        r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=last_col)
    c = ws.cell(row=r, column=2, value=f"{statement_title} - {fy_label}")
    c.font = _tnr(11, bold=True)
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    r += 2
    return r


def _header_font():
    return Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)


def _title_font():
    return Font(name=FONT_NAME, bold=True, size=14, color="1F4E79")


def _bold():
    return Font(name=FONT_NAME, bold=True, size=10)


def _normal():
    return Font(name=FONT_NAME, size=10)


def _write_title_block(ws, company: Company, statement_title: str, fy_label: str, next_row: int = 1):
    ws.cell(row=next_row, column=1, value=company.name).font = _title_font()
    ws.cell(row=next_row + 1, column=1, value=statement_title).font = Font(name=FONT_NAME, bold=True, size=12)
    ws.cell(row=next_row + 2, column=1, value=f"For the year ended - {fy_label}").font = Font(
        name=FONT_NAME, italic=True, size=10)
    ws.cell(row=next_row + 3, column=1, value=f"(All amounts in {company.currency} unless otherwise stated)").font = \
        Font(name=FONT_NAME, italic=True, size=9)
    return next_row + 5


def _autosize(ws, widths: dict[int, int]):
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def build_workbook(
    company: Company,
    fy_label: str,
    tb: TrialBalance,
    mappings: dict[str, MappingEntry],
    bs: BalanceSheet,
    pnl: ProfitAndLoss,
    cash_flow: CashFlowStatement,
    notes: list[Note],
    issues: list[ValidationIssue],
    ratios: list[RatioResult],
    soce: StatementOfChangesInEquity | None = None,
    receivables_ageing: AgeingGrid | None = None,
    payables_ageing: AgeingGrid | None = None,
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    _write_tb_sheet(wb, company, fy_label, tb, mappings)
    _write_mapping_sheet(wb, company, fy_label, tb, mappings)
    note_rows = _write_notes_sheet(wb, company, fy_label, notes)
    _write_balance_sheet_sheet(wb, company, fy_label, bs, note_rows)
    _write_pnl_sheet(wb, company, fy_label, pnl, note_rows)
    if soce is not None:
        _write_soce_sheet(wb, company, fy_label, soce)
    _write_cash_flow_sheet(wb, company, fy_label, cash_flow)
    if receivables_ageing is not None or payables_ageing is not None:
        _write_ageing_sheet(wb, company, fy_label, receivables_ageing, payables_ageing)
    _write_validation_sheet(wb, company, fy_label, issues)
    _write_ratios_sheet(wb, company, fy_label, ratios)

    # Tab order: Notes was written first (so its row numbers exist for BS/P&L
    # formulas) but reads better placed right after P&L for a reviewer.
    order = wb.sheetnames
    order.remove("Notes")
    pnl_idx = order.index("P&L")
    order.insert(pnl_idx + 1, "Notes")
    wb._sheets = [wb[name] for name in order]

    return wb


def _write_soce_sheet(wb, company, fy_label, soce: StatementOfChangesInEquity):
    ws = wb.create_sheet("SOCE")
    r = _write_title_block(ws, company, "Statement of Changes in Equity", fy_label)

    ws.cell(row=r, column=1, value="A. Equity Share Capital").font = Font(
        name=FONT_NAME, bold=True, color="1F4E79")
    r += 1
    sc = soce.equity_share_capital
    for label, val in [
        ("Balance at the beginning of the year", sc.opening),
        ("Changes in equity share capital during the year", sc.changes_during_year),
        ("Balance at the end of the year", sc.closing),
    ]:
        ws.cell(row=r, column=1, value=label).font = _normal()
        ws.cell(row=r, column=2, value=val).number_format = NUM_FMT
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="B. Other Equity").font = Font(
        name=FONT_NAME, bold=True, color="1F4E79")
    r += 1
    headers = ["Component", "Opening Balance", "Profit for the Year", "Other Additions/(Deductions)", "Closing Balance"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = _header_font()
        cell.fill = HEADER_FILL
    r += 1
    for comp in soce.other_equity:
        ws.cell(row=r, column=1, value=comp.component).font = _normal()
        ws.cell(row=r, column=2, value=comp.opening).number_format = NUM_FMT
        ws.cell(row=r, column=3, value=comp.profit_for_the_year).number_format = NUM_FMT
        ws.cell(row=r, column=4, value=comp.other_movements).number_format = NUM_FMT
        ws.cell(row=r, column=5, value=comp.closing).number_format = NUM_FMT
        r += 1
    ws.cell(row=r, column=1, value="Total Other Equity").font = _bold()
    ws.cell(row=r, column=2, value=soce.total_other_equity_opening).number_format = NUM_FMT
    ws.cell(row=r, column=5, value=soce.total_other_equity_closing).number_format = NUM_FMT
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
    r += 2

    ws.cell(row=r, column=1, value=(
        "Note: \"Other Additions/(Deductions)\" is a residual figure (closing "
        "less opening less profit transferred) capturing dividends paid, "
        "transfers between reserves, or prior period adjustments -- a Trial "
        "Balance alone cannot distinguish these; verify against board "
        "resolutions / minutes before finalizing."
    )).font = Font(name=FONT_NAME, italic=True, size=9, color="808080")
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    _autosize(ws, {1: 40, 2: 18, 3: 18, 4: 24, 5: 18})
    ws.sheet_view.showGridLines = False


def _write_ageing_sheet(wb, company, fy_label, receivables: AgeingGrid | None, payables: AgeingGrid | None):
    ws = wb.create_sheet("Ageing Schedule")
    r = _write_title_block(ws, company, "Ageing Schedule - Trade Receivables and Trade Payables", fy_label)

    def write_grid(grid: AgeingGrid, title: str, row: int) -> int:
        ws.cell(row=row, column=1, value=title).font = Font(name=FONT_NAME, bold=True, color="1F4E79")
        row += 1
        if not grid.available:
            ws.cell(row=row, column=1, value=(
                f"Ageing schedule not available: {grid.unavailable_reason}"
            )).font = Font(name=FONT_NAME, italic=True, size=10, color="C00000")
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            return row + 3

        headers = ["Category"] + BUCKET_LABELS + ["Total"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = _header_font()
            cell.fill = HEADER_FILL
        row += 1
        for cat in CATEGORY_LABELS:
            ws.cell(row=row, column=1, value=cat).font = _normal()
            row_total = 0.0
            for c, bucket in enumerate(BUCKET_LABELS, start=2):
                val = grid.grid[cat][bucket]
                row_total += val
                ws.cell(row=row, column=c, value=val).number_format = NUM_FMT
            ws.cell(row=row, column=len(BUCKET_LABELS) + 2, value=round(row_total, 2)).number_format = NUM_FMT
            row += 1
        ws.cell(row=row, column=1, value="Total").font = _bold()
        ws.cell(row=row, column=len(BUCKET_LABELS) + 2, value=grid.total).number_format = NUM_FMT
        for c in range(1, len(BUCKET_LABELS) + 3):
            ws.cell(row=row, column=c).fill = TOTAL_FILL
        row += 1
        if grid.reconciles_to_balance_sheet is False:
            ws.cell(row=row, column=1, value=(
                f"WARNING: Ageing total ({grid.total:,.2f}) does not reconcile to the "
                f"Balance Sheet figure ({grid.balance_sheet_amount:,.2f})."
            )).font = Font(name=FONT_NAME, color="C00000", bold=True)
            row += 1
        return row + 2

    if receivables is not None:
        r = write_grid(receivables, "Trade Receivables Ageing (as of Balance Sheet date)", r)
    if payables is not None:
        r = write_grid(payables, "Trade Payables Ageing (as of Balance Sheet date)", r)

    _autosize(ws, {1: 32, 2: 14, 3: 14, 4: 14, 5: 12, 6: 12, 7: 14, 8: 16})
    ws.sheet_view.showGridLines = False


def _write_tb_sheet(wb, company, fy_label, tb: TrialBalance, mappings):
    ws = wb.create_sheet("TB")
    row = _write_title_block(ws, company, "Trial Balance", fy_label)
    headers = ["Ledger Name", "Opening Balance", "Debit", "Credit", "Closing Balance", "Mapped Head"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = _header_font()
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    start_data_row = row + 1
    r = start_data_row
    for ledger in tb.ledgers:
        mapping = mappings.get(ledger.ledger_name)
        ws.cell(row=r, column=1, value=ledger.ledger_name).font = _normal()
        ws.cell(row=r, column=2, value=ledger.opening_balance).number_format = NUM_FMT
        ws.cell(row=r, column=3, value=ledger.debit).number_format = NUM_FMT
        ws.cell(row=r, column=4, value=ledger.credit).number_format = NUM_FMT
        ws.cell(row=r, column=5, value=f"=B{r}+C{r}-D{r}").number_format = NUM_FMT
        ws.cell(row=r, column=6, value=f"{mapping.major_head} / {mapping.sub_head}" if mapping else "UNMAPPED")
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = BORDER
        r += 1
    total_row = r
    ws.cell(row=total_row, column=1, value="Total").font = _bold()
    for c, col_letter in [(2, "B"), (3, "C"), (4, "D")]:
        ws.cell(row=total_row, column=c,
                value=f"=SUM({col_letter}{start_data_row}:{col_letter}{total_row - 1})").font = _bold()
        ws.cell(row=total_row, column=c).number_format = NUM_FMT
        ws.cell(row=total_row, column=c).fill = TOTAL_FILL
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    _autosize(ws, {1: 38, 2: 16, 3: 16, 4: 16, 5: 16, 6: 42})
    ws.freeze_panes = f"A{start_data_row}"


def _write_mapping_sheet(wb, company, fy_label, tb, mappings):
    ws = wb.create_sheet("Mapping")
    row = _write_title_block(ws, company, "Ledger Mapping (Editable)", fy_label)
    headers = ["Ledger Name", "Closing Balance", "Major Head", "Sub Head", "Current / Non-Current",
               "Nature", "Note Reference", "Source", "Confidence"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = _header_font()
        cell.fill = HEADER_FILL
        cell.border = BORDER
    r = row + 1
    for ledger in tb.ledgers:
        m = mappings.get(ledger.ledger_name)
        ws.cell(row=r, column=1, value=ledger.ledger_name)
        ws.cell(row=r, column=2, value=ledger.closing_balance).number_format = NUM_FMT
        ws.cell(row=r, column=3, value=m.major_head if m else "")
        ws.cell(row=r, column=4, value=m.sub_head if m else "PENDING - PLEASE MAP")
        ws.cell(row=r, column=5, value=m.current_or_non_current.value if m else "")
        ws.cell(row=r, column=6, value=m.nature.value if m else "")
        ws.cell(row=r, column=7, value=m.note_ref if m else "")
        ws.cell(row=r, column=8, value=m.source if m else "")
        ws.cell(row=r, column=9, value=m.confidence if m else 0)
        if not m:
            for c in range(1, 10):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FFF2CC")
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = _normal()
        r += 1
    _autosize(ws, {1: 38, 2: 16, 3: 24, 4: 30, 5: 18, 6: 12, 7: 30, 8: 14, 9: 12})
    ws.freeze_panes = f"A{row + 1}"


def _write_section(ws, r, title, majors_or_subs, is_major_list: bool):
    ws.cell(row=r, column=1, value=title).font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    r += 1
    grand_cy = grand_py = 0.0
    for entry in majors_or_subs:
        ws.cell(row=r, column=1, value=entry.major_head if is_major_list else entry.sub_head).font = _bold()
        ws.cell(row=r, column=1).fill = SUBHEAD_FILL
        ws.cell(row=r, column=3, value=entry.current_year).number_format = NUM_FMT
        ws.cell(row=r, column=4, value=entry.previous_year).number_format = NUM_FMT
        ws.cell(row=r, column=3).font = _bold()
        ws.cell(row=r, column=4).font = _bold()
        ws.cell(row=r, column=1).fill = SUBHEAD_FILL
        ws.cell(row=r, column=3).fill = SUBHEAD_FILL
        ws.cell(row=r, column=4).fill = SUBHEAD_FILL
        r += 1
        grand_cy += entry.current_year
        grand_py += entry.previous_year
        if is_major_list:
            for sh in entry.sub_heads:
                ws.cell(row=r, column=2, value=sh.sub_head).font = _normal()
                ws.cell(row=r, column=3, value=sh.current_year).number_format = NUM_FMT
                ws.cell(row=r, column=4, value=sh.previous_year).number_format = NUM_FMT
                ws.cell(row=r, column=5, value=sh.note_ref or "").font = Font(name=FONT_NAME, size=9, italic=True)
                r += 1
    return r, grand_cy, grand_py


def _period_labels(company: Company) -> tuple[str, str]:
    cy_end = company.financial_year_end
    try:
        py_end = cy_end.replace(year=cy_end.year - 1)
    except ValueError:
        py_end = cy_end.replace(year=cy_end.year - 1, day=28)
    fmt = "%d %B %Y"
    return cy_end.strftime(fmt), py_end.strftime(fmt)


def _bs_note_value(ws, r, col_cy, col_py, sh, note_rows: dict[str, int]):
    """Write a sub-head's CY/PY figures as a live formula to its Notes-sheet
    total when a note exists, falling back to the computed value otherwise."""
    note_row = note_rows.get(sh.note_ref or "")
    if note_row:
        ws.cell(row=r, column=col_cy, value=f"='Notes'!F{note_row}")
        ws.cell(row=r, column=col_py, value=f"='Notes'!G{note_row}")
    else:
        ws.cell(row=r, column=col_cy, value=sh.current_year)
        ws.cell(row=r, column=col_py, value=sh.previous_year)
    ws.cell(row=r, column=col_cy).number_format = ACCT_FMT
    ws.cell(row=r, column=col_py).number_format = ACCT_FMT


def _write_signature_block(ws, company: Company, r: int) -> int:
    ws.cell(row=r, column=3, value="See accompanying notes forming part of the financial statements").font = \
        _tnr(9, italic=True)
    r += 2
    ws.cell(row=r, column=2, value="For and on behalf of the Board of Directors").font = _tnr(10, bold=True)
    ws.cell(row=r, column=7, value="As per our report of even date attached").font = _tnr(10, bold=True)
    r += 1
    ws.cell(row=r, column=7, value=f"For {company.auditor or '<Auditor Firm Name>'}").font = _tnr(10)
    r += 1
    ws.cell(row=r, column=7, value="Chartered Accountants").font = _tnr(10)
    r += 1
    ws.cell(row=r, column=7, value="Firm Registration No.:").font = _tnr(10)
    r += 4
    d1 = company.directors[0] if len(company.directors) > 0 else "Director"
    d2 = company.directors[1] if len(company.directors) > 1 else "Director"
    ws.cell(row=r, column=2, value=d1).font = _tnr(10)
    ws.cell(row=r, column=3, value=d2).font = _tnr(10)
    ws.cell(row=r, column=7, value="Partner").font = _tnr(10)
    r += 1
    ws.cell(row=r, column=2, value="Director").font = _tnr(9, italic=True)
    ws.cell(row=r, column=3, value="Director").font = _tnr(9, italic=True)
    ws.cell(row=r, column=7, value="Membership No.:").font = _tnr(10)
    r += 1
    ws.cell(row=r, column=2, value="DIN:").font = _tnr(10)
    ws.cell(row=r, column=3, value="DIN:").font = _tnr(10)
    r += 3
    ws.cell(row=r, column=2, value="Place:").font = _tnr(10)
    r += 1
    ws.cell(row=r, column=2, value="Date:").font = _tnr(10)
    r += 1
    ws.cell(row=r, column=2, value="UDIN:").font = _tnr(10)
    return r


def _write_balance_sheet_sheet(wb, company, fy_label, bs: BalanceSheet, note_rows: dict[str, int]):
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False
    cy_label, py_label = _period_labels(company)
    r = _write_schedule3_header(ws, company, "Balance Sheet", fy_label, last_col=8)

    hdr_row = r
    ws.cell(row=r, column=2, value="Sl. No.")
    ws.cell(row=r, column=3, value="Particulars")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    ws.cell(row=r, column=6, value="Note No.")
    ws.cell(row=r, column=7, value=f"As at\n{cy_label}\nAmount (Rs.)")
    ws.cell(row=r, column=8, value=f"As at\n{py_label}\nAmount (Rs.)")
    for c in (2, 3, 6, 7, 8):
        cell = ws.cell(row=r, column=c)
        cell.font = _tnr(10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 42
    _box_border(ws, r, 2, r, 8, "medium")
    r += 2

    def write_group(major_list, group_numeral, group_title, start_row):
        rr = start_row
        ws.cell(row=rr, column=2, value=group_numeral).font = _tnr(11, bold=True)
        ws.cell(row=rr, column=3, value=group_title).font = _tnr(11, bold=True)
        rr += 1
        subtotal_rows = []
        for idx, major in enumerate(major_list, start=1):
            ws.cell(row=rr, column=3, value=f"({idx}) {major.major_head}").font = _tnr(10, bold=True)
            rr += 1
            sub_start = rr
            for s_idx, sh in enumerate(major.sub_heads):
                letter = chr(ord('a') + s_idx)
                ws.cell(row=rr, column=3, value=f"     ({letter}) {sh.sub_head}").font = _tnr(10)
                ws.cell(row=rr, column=6, value=_note_number(sh.note_ref)).alignment = Alignment(horizontal="center")
                ws.cell(row=rr, column=6).font = _tnr(10)
                _bs_note_value(ws, rr, 7, 8, sh, note_rows)
                rr += 1
            sub_end = rr - 1
            ws.cell(row=rr, column=3, value=f"Total - {major.major_head}").font = _tnr(10, bold=True)
            ws.cell(row=rr, column=7, value=f"=SUM(G{sub_start}:G{sub_end})").font = _tnr(10, bold=True)
            ws.cell(row=rr, column=8, value=f"=SUM(H{sub_start}:H{sub_end})").font = _tnr(10, bold=True)
            ws.cell(row=rr, column=7).number_format = ACCT_FMT
            ws.cell(row=rr, column=8).number_format = ACCT_FMT
            _box_border(ws, rr, 7, rr, 8, "thin")
            subtotal_rows.append(rr)
            rr += 2
        return rr, subtotal_rows

    r, el_subtotals = write_group(bs.equity_and_liabilities, "I", "EQUITY AND LIABILITIES", r)
    ws.cell(row=r, column=4, value="TOTAL - EQUITY AND LIABILITIES").font = _tnr(10, bold=True)
    ws.cell(row=r, column=7, value="=" + "+".join(f"G{rr}" for rr in el_subtotals)).font = _tnr(10, bold=True)
    ws.cell(row=r, column=8, value="=" + "+".join(f"H{rr}" for rr in el_subtotals)).font = _tnr(10, bold=True)
    for c in (7, 8):
        ws.cell(row=r, column=c).number_format = ACCT_FMT
        ws.cell(row=r, column=c).fill = TOTAL_FILL
    ws.cell(row=r, column=4).fill = TOTAL_FILL
    _box_border(ws, r, 7, r, 8, "double")
    el_total_row = r
    r += 3

    r, a_subtotals = write_group(bs.assets, "II", "ASSETS", r)
    ws.cell(row=r, column=4, value="TOTAL - ASSETS").font = _tnr(10, bold=True)
    ws.cell(row=r, column=7, value="=" + "+".join(f"G{rr}" for rr in a_subtotals)).font = _tnr(10, bold=True)
    ws.cell(row=r, column=8, value="=" + "+".join(f"H{rr}" for rr in a_subtotals)).font = _tnr(10, bold=True)
    for c in (7, 8):
        ws.cell(row=r, column=c).number_format = ACCT_FMT
        ws.cell(row=r, column=c).fill = TOTAL_FILL
    ws.cell(row=r, column=4).fill = TOTAL_FILL
    _box_border(ws, r, 7, r, 8, "double")
    as_total_row = r
    r += 2

    ws.cell(row=r, column=3, value="Difference (Equity & Liabilities - Assets), should be Nil").font = \
        _tnr(9, italic=True, color="808080")
    ws.cell(row=r, column=7, value=f"=G{el_total_row}-G{as_total_row}").font = _tnr(9, italic=True, color="808080")
    ws.cell(row=r, column=7).number_format = ACCT_FMT
    r += 2

    r = _write_signature_block(ws, company, r)
    _autosize(ws, {1: 3, 2: 7, 3: 34, 4: 16, 5: 10, 6: 9, 7: 20, 8: 20})
    _fit_to_page(ws, last_col=8)
    ws.print_title_rows = f"{hdr_row}:{hdr_row}"


def _write_pnl_sheet(wb, company, fy_label, pnl: ProfitAndLoss, note_rows: dict[str, int]):
    ws = wb.create_sheet("P&L")
    ws.sheet_view.showGridLines = False
    cy_label, py_label = _period_labels(company)
    r = _write_schedule3_header(ws, company, "Statement of Profit and Loss", fy_label, last_col=8)

    hdr_row = r
    ws.cell(row=r, column=2, value="Sl. No.")
    ws.cell(row=r, column=3, value="Particulars")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    ws.cell(row=r, column=6, value="Note No.")
    ws.cell(row=r, column=7, value=f"For the year ended\n{cy_label}")
    ws.cell(row=r, column=8, value=f"For the year ended\n{py_label}")
    for c in (2, 3, 6, 7, 8):
        cell = ws.cell(row=r, column=c)
        cell.font = _tnr(10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 32
    _box_border(ws, r, 2, r, 8, "medium")
    r += 2

    ws.cell(row=r, column=2, value="I").font = _tnr(11, bold=True)
    ws.cell(row=r, column=3, value="Revenue").font = _tnr(11, bold=True)
    r += 1
    rev_start = r
    for sh in pnl.revenue:
        ws.cell(row=r, column=3, value=sh.sub_head).font = _tnr(10)
        ws.cell(row=r, column=6, value=_note_number(sh.note_ref)).alignment = Alignment(horizontal="center")
        _bs_note_value(ws, r, 7, 8, sh, note_rows)
        r += 1
    rev_end = r - 1
    ws.cell(row=r, column=3, value="Total Revenue").font = _tnr(10, bold=True)
    ws.cell(row=r, column=7, value=f"=SUM(G{rev_start}:G{rev_end})").font = _tnr(10, bold=True)
    ws.cell(row=r, column=8, value=f"=SUM(H{rev_start}:H{rev_end})").font = _tnr(10, bold=True)
    for c in (7, 8):
        ws.cell(row=r, column=c).number_format = ACCT_FMT
        ws.cell(row=r, column=c).fill = TOTAL_FILL
    _box_border(ws, r, 7, r, 8, "thin")
    total_rev_row = r
    r += 2

    ws.cell(row=r, column=2, value="II").font = _tnr(11, bold=True)
    ws.cell(row=r, column=3, value="Expenses").font = _tnr(11, bold=True)
    r += 1
    exp_start = r
    for sh in pnl.expenses:
        ws.cell(row=r, column=3, value=sh.sub_head).font = _tnr(10)
        ws.cell(row=r, column=6, value=_note_number(sh.note_ref)).alignment = Alignment(horizontal="center")
        _bs_note_value(ws, r, 7, 8, sh, note_rows)
        r += 1
    exp_end = r - 1
    ws.cell(row=r, column=3, value="Total Expenses").font = _tnr(10, bold=True)
    ws.cell(row=r, column=7, value=f"=SUM(G{exp_start}:G{exp_end})").font = _tnr(10, bold=True)
    ws.cell(row=r, column=8, value=f"=SUM(H{exp_start}:H{exp_end})").font = _tnr(10, bold=True)
    for c in (7, 8):
        ws.cell(row=r, column=c).number_format = ACCT_FMT
        ws.cell(row=r, column=c).fill = TOTAL_FILL
    _box_border(ws, r, 7, r, 8, "thin")
    total_exp_row = r
    r += 2

    ws.cell(row=r, column=2, value="III").font = _tnr(11, bold=True)
    ws.cell(row=r, column=3, value="Profit Before Tax (I - II)").font = _tnr(10, bold=True)
    ws.cell(row=r, column=7, value=f"=G{total_rev_row}-G{total_exp_row}").font = _tnr(10, bold=True)
    ws.cell(row=r, column=8, value=f"=H{total_rev_row}-H{total_exp_row}").font = _tnr(10, bold=True)
    for c in (7, 8):
        ws.cell(row=r, column=c).number_format = ACCT_FMT
    pbt_row = r
    r += 2

    ws.cell(row=r, column=2, value="IV").font = _tnr(11, bold=True)
    ws.cell(row=r, column=3, value="Tax Expense").font = _tnr(10, bold=True)
    ws.cell(row=r, column=7, value=pnl.tax_expense_cy).number_format = ACCT_FMT
    ws.cell(row=r, column=8, value=pnl.tax_expense_py).number_format = ACCT_FMT
    tax_row = r
    r += 2

    ws.cell(row=r, column=2, value="V").font = _tnr(11, bold=True)
    ws.cell(row=r, column=3, value="Profit After Tax (III - IV)").font = _tnr(10, bold=True)
    ws.cell(row=r, column=7, value=f"=G{pbt_row}-G{tax_row}").font = _tnr(10, bold=True)
    ws.cell(row=r, column=8, value=f"=H{pbt_row}-H{tax_row}").font = _tnr(10, bold=True)
    for c in (7, 8):
        ws.cell(row=r, column=c).number_format = ACCT_FMT
        ws.cell(row=r, column=c).fill = TOTAL_FILL
    _box_border(ws, r, 7, r, 8, "double")
    r += 2

    r = _write_signature_block(ws, company, r)
    _autosize(ws, {1: 3, 2: 7, 3: 34, 4: 16, 5: 10, 6: 9, 7: 20, 8: 20})
    _fit_to_page(ws, last_col=8)
    ws.print_title_rows = f"{hdr_row}:{hdr_row}"


def _write_cash_flow_sheet(wb, company, fy_label, cf: CashFlowStatement):
    ws = wb.create_sheet("Cash Flow")
    r = _write_title_block(ws, company, "Cash Flow Statement (Indirect Method)", fy_label)
    ws.cell(row=r, column=1, value="A. Cash Flow from Operating Activities").font = Font(
        name=FONT_NAME, bold=True, color="1F4E79")
    r += 1
    for label, val in [
        ("Net Profit Before Tax", cf.net_profit_before_tax),
        ("Add: Depreciation", cf.depreciation_addback),
        ("Add: Interest Expense", cf.interest_expense_addback),
    ]:
        ws.cell(row=r, column=1, value=label).font = _normal()
        ws.cell(row=r, column=3, value=val).number_format = NUM_FMT
        r += 1
    ws.cell(row=r, column=1, value="Working Capital Adjustments:").font = _bold()
    r += 1
    for label, val in cf.working_capital_changes.items():
        ws.cell(row=r, column=2, value=f"(Increase)/Decrease in {label}").font = _normal()
        ws.cell(row=r, column=3, value=val).number_format = NUM_FMT
        r += 1
    ws.cell(row=r, column=1, value="Net Cash from Operating Activities").font = _bold()
    ws.cell(row=r, column=3, value=cf.cash_from_operations).number_format = NUM_FMT
    ws.cell(row=r, column=1).fill = TOTAL_FILL
    ws.cell(row=r, column=3).fill = TOTAL_FILL
    r += 2

    ws.cell(row=r, column=1, value="B. Cash Flow from Investing Activities").font = Font(
        name=FONT_NAME, bold=True, color="1F4E79")
    r += 1
    ws.cell(row=r, column=1, value="Net Cash used in Investing Activities").font = _bold()
    ws.cell(row=r, column=3, value=cf.cash_from_investing).number_format = NUM_FMT
    r += 2

    ws.cell(row=r, column=1, value="C. Cash Flow from Financing Activities").font = Font(
        name=FONT_NAME, bold=True, color="1F4E79")
    r += 1
    ws.cell(row=r, column=1, value="Net Cash from/used in Financing Activities").font = _bold()
    ws.cell(row=r, column=3, value=cf.cash_from_financing).number_format = NUM_FMT
    r += 2

    for label, val in [
        ("Net Increase/(Decrease) in Cash and Cash Equivalents (A+B+C)", cf.net_increase_in_cash),
        ("Cash and Cash Equivalents at the Beginning of the Year", cf.opening_cash),
        ("Cash and Cash Equivalents at the End of the Year", cf.closing_cash),
    ]:
        ws.cell(row=r, column=1, value=label).font = _bold()
        ws.cell(row=r, column=3, value=val).number_format = NUM_FMT
        r += 1

    r += 1
    ws.cell(row=r, column=1, value=(
        "Note: Working capital movements require two years of Trial Balance "
        "data (previous year closing balances) to be populated for accurate "
        "results. Movements shown as 0 indicate previous-year data was not "
        "supplied for that ledger."
    )).font = Font(name=FONT_NAME, italic=True, size=9, color="808080")
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    _autosize(ws, {1: 46, 2: 40, 3: 18})
    ws.sheet_view.showGridLines = False


def _write_notes_sheet(wb, company, fy_label, notes: list[Note]) -> dict[str, int]:
    ws = wb.create_sheet("Notes")
    ws.sheet_view.showGridLines = False
    cy_label, py_label = _period_labels(company)
    r = _write_schedule3_header(ws, company, "Notes forming part of the Financial Statements", fy_label, last_col=7)
    note_rows: dict[str, int] = {}
    for note in notes:
        block_start = r
        ws.cell(row=r, column=2, value=note.note_ref).font = _tnr(11, bold=True, color="FFFFFF")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        ws.cell(row=r, column=2).fill = HEADER_FILL
        r += 1
        ws.cell(row=r, column=2, value="Particulars").font = _tnr(10, bold=True)
        ws.cell(row=r, column=6, value=f"As at / Year ended\n{cy_label}").font = _tnr(9, bold=True)
        ws.cell(row=r, column=7, value=f"As at / Year ended\n{py_label}").font = _tnr(9, bold=True)
        for c in (6, 7):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", wrap_text=True)
        r += 1
        item_start = r
        for item in note.line_items:
            ws.cell(row=r, column=2, value=item.label).font = _tnr(10)
            ws.cell(row=r, column=6, value=item.current_year).number_format = ACCT_FMT
            ws.cell(row=r, column=7, value=item.previous_year).number_format = ACCT_FMT
            r += 1
        item_end = r - 1
        ws.cell(row=r, column=2, value="Total").font = _tnr(10, bold=True)
        if item_end >= item_start:
            ws.cell(row=r, column=6, value=f"=SUM(F{item_start}:F{item_end})").font = _tnr(10, bold=True)
            ws.cell(row=r, column=7, value=f"=SUM(G{item_start}:G{item_end})").font = _tnr(10, bold=True)
        else:
            ws.cell(row=r, column=6, value=note.total_current_year).font = _tnr(10, bold=True)
            ws.cell(row=r, column=7, value=note.total_previous_year).font = _tnr(10, bold=True)
        ws.cell(row=r, column=6).number_format = ACCT_FMT
        ws.cell(row=r, column=7).number_format = ACCT_FMT
        ws.cell(row=r, column=2).fill = TOTAL_FILL
        ws.cell(row=r, column=6).fill = TOTAL_FILL
        ws.cell(row=r, column=7).fill = TOTAL_FILL
        _box_border(ws, block_start, 2, r, 7, "thin")
        _box_border(ws, r, 6, r, 7, "double")
        note_rows[note.note_ref] = r
        r += 2
    _autosize(ws, {1: 3, 2: 44, 3: 10, 4: 10, 5: 10, 6: 20, 7: 20})
    _fit_to_page(ws, last_col=7)
    return note_rows


def _write_validation_sheet(wb, company, fy_label, issues: list[ValidationIssue]):
    ws = wb.create_sheet("Validation")
    r = _write_title_block(ws, company, "Validation Report", fy_label)
    headers = ["Severity", "Code", "Ledger", "Message", "Amount"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = _header_font()
        cell.fill = HEADER_FILL
    r += 1
    severity_fill = {"ERROR": "F8CBAD", "WARNING": "FFE699", "INFO": "D9E1F2"}
    for issue in issues:
        ws.cell(row=r, column=1, value=issue.severity)
        ws.cell(row=r, column=2, value=issue.code)
        ws.cell(row=r, column=3, value=issue.ledger_name or "")
        ws.cell(row=r, column=4, value=issue.message).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=5, value=issue.amount if issue.amount is not None else "")
        fill = PatternFill("solid", fgColor=severity_fill.get(issue.severity, "FFFFFF"))
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).font = _normal()
        r += 1
    if not issues:
        ws.cell(row=r, column=1, value="No issues found.").font = Font(name=FONT_NAME, color="006100", bold=True)
    _autosize(ws, {1: 12, 2: 22, 3: 30, 4: 70, 5: 16})
    ws.sheet_view.showGridLines = False


def _write_ratios_sheet(wb, company, fy_label, ratios: list[RatioResult]):
    ws = wb.create_sheet("Ratios")
    r = _write_title_block(ws, company, "Financial Ratio Analysis", fy_label)
    headers = ["Ratio", "Current Year", "Previous Year", "Formula"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = _header_font()
        cell.fill = HEADER_FILL
    r += 1
    for ratio in ratios:
        ws.cell(row=r, column=1, value=ratio.name).font = _normal()
        ws.cell(row=r, column=2, value=ratio.current_year if ratio.current_year is not None else "N/A")
        ws.cell(row=r, column=3, value=ratio.previous_year if ratio.previous_year is not None else "N/A")
        ws.cell(row=r, column=4, value=ratio.formula).font = Font(name=FONT_NAME, italic=True, size=9)
        r += 1
    _autosize(ws, {1: 32, 2: 16, 3: 16, 4: 50})
    ws.sheet_view.showGridLines = False


def save_workbook(wb: Workbook, output_path: str) -> str:
    wb.save(output_path)
    return output_path
